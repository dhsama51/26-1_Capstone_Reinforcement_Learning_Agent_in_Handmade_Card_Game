from __future__ import annotations

import json
import re
import zipfile
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook
from openpyxl.chart.text import Paragraph, RichText
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.text import CharacterProperties, ParagraphProperties, RegularTextRun
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parent
ZIP_PATH = ROOT / "log" / "make_balance_after_patch2_with_nnn_four_agent.zip"
OUT_PATH = ROOT / "log" / "make_balance_after_patch2_with_nnn_four_agent_hp6_hist_formatted_v2.xlsx"
TEMPLATE_PATH = Path(r"C:\Users\user\Downloads\history_match_results_deck_400_formula.xlsx")

KOR_DECK = {"Orange": "귤", "Charlotte": "샤를로테"}
DECK_MAX_HP = {"Orange": 6, "Charlotte": 10}
HP_BINS = list(range(-2, 11))

SCENARIO_RE = re.compile(
    r"^(?P<idx>\d+)_(?P<side>P1|P2)_(?P<self_deck>Orange|Charlotte)_vs_(?P<opp_deck>Orange|Charlotte)__"
)
MATCH_RE = re.compile(
    r"- match (?P<match_no>\d+): result=(?P<result>\w+) steps=(?P<steps>\d+) turn=(?P<turn>\d+) game_id=(?P<game_id>\S+)"
)
LEADER_JSON_RE = re.compile(r"leader_hp=(\{.*\})")


@dataclass
class ScenarioMeta:
    idx: int
    side: str
    self_deck: str
    opp_deck: str
    source_name: str

    @property
    def self_side_kor(self) -> str:
        return "선공" if self.side == "P1" else "후공"

    @property
    def relation_kor(self) -> str:
        return "같은 덱" if self.self_deck == self.opp_deck else "다른 덱"

    @property
    def self_deck_kor(self) -> str:
        return KOR_DECK[self.self_deck]

    @property
    def opp_deck_kor(self) -> str:
        return KOR_DECK[self.opp_deck]


def _parse_scenario_meta(path: Path) -> ScenarioMeta:
    m = SCENARIO_RE.match(path.name)
    if not m:
        raise ValueError(f"Unexpected scenario filename: {path.name}")
    return ScenarioMeta(
        idx=int(m.group("idx")),
        side=m.group("side"),
        self_deck=m.group("self_deck"),
        opp_deck=m.group("opp_deck"),
        source_name=path.name,
    )


def _iter_hist_files(zf: zipfile.ZipFile) -> List[str]:
    names = []
    for name in zf.namelist():
        if name.endswith("_hist.txt"):
            names.append(name)
    return sorted(names)


def _parse_hist_text(text: str) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    current_match: Dict[str, object] | None = None

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        m = MATCH_RE.search(line)
        if m:
            current_match = {
                "match_no": int(m.group("match_no")),
                "result": m.group("result"),
                "steps": int(m.group("steps")),
                "turn": int(m.group("turn")),
                "game_id": m.group("game_id"),
            }
            continue
        if current_match is not None:
            mj = LEADER_JSON_RE.search(line)
            if mj:
                payload = json.loads(mj.group(1))
                current_match["leader_hp"] = payload
                rows.append(current_match)
                current_match = None

    return rows


def _player_row(leader_hp: Dict[str, object], player_key: str) -> Dict[str, object]:
    entry = leader_hp["leader_hp_by_player"][player_key]
    return {
        "deck": str(entry["deck"]),
        "final_hp": int(round(float(entry["final_hp"]))),
        "min_hp": int(round(float(entry["min_hp"]))),
    }


def _match_rows_from_file(zf: zipfile.ZipFile, name: str) -> List[Dict[str, object]]:
    meta = _parse_scenario_meta(Path(name))
    text = zf.read(name).decode("utf-8", errors="replace")
    parsed = _parse_hist_text(text)

    rows: List[Dict[str, object]] = []
    for item in parsed:
        leader_hp = item["leader_hp"]
        p1 = _player_row(leader_hp, "P1")
        p2 = _player_row(leader_hp, "P2")
        if meta.side == "P1":
            self_player = p1
            opp_player = p2
            p1_deck = meta.self_deck
            p2_deck = meta.opp_deck
        else:
            self_player = p2
            opp_player = p1
            p1_deck = meta.opp_deck
            p2_deck = meta.self_deck

        self_max = DECK_MAX_HP[meta.self_deck]
        hp_check = "OK" if self_max == DECK_MAX_HP[self_player["deck"]] else "CHECK"
        self_wl = {
            "Player1Win": "승" if meta.side == "P1" else "패",
            "Player2Win": "패" if meta.side == "P1" else "승",
            "Draw": "무",
        }.get(str(item["result"]), "무")

        rows.append(
            {
                "SourceFile": meta.source_name,
                "MatchNo": int(item["match_no"]),
                "GameID": str(item["game_id"]),
                "SelfSide": meta.self_side_kor,
                "SelfDeck": meta.self_deck_kor,
                "OppDeck": meta.opp_deck_kor,
                "Relation": meta.relation_kor,
                "Result": str(item["result"]),
                "Self W/L": self_wl,
                "Steps": int(item["steps"]),
                "FinalTurn": int(item["turn"]),
                "P1Deck": KOR_DECK[p1_deck],
                "P1FinalHP": p1["final_hp"],
                "P1MinHP": p1["min_hp"],
                "P2Deck": KOR_DECK[p2_deck],
                "P2FinalHP": p2["final_hp"],
                "P2MinHP": p2["min_hp"],
                "SelfFinalHP": self_player["final_hp"],
                "SelfMinHP": self_player["min_hp"],
                "OppFinalHP": opp_player["final_hp"],
                "OppMinHP": opp_player["min_hp"],
                "SelfMaxHP": self_max,
                "HP_Check": hp_check,
            }
        )
    return rows


def _build_match_data(zf: zipfile.ZipFile) -> List[Dict[str, object]]:
    all_rows: List[Dict[str, object]] = []
    for name in _iter_hist_files(zf):
        all_rows.extend(_match_rows_from_file(zf, name))
    return all_rows


def _bins_for_deck(rows: Iterable[Dict[str, object]], deck_label: str) -> Dict[str, Counter]:
    min_counts = Counter()
    final_counts = Counter()
    for row in rows:
        if row["SelfDeck"] != deck_label:
            continue
        min_counts[int(row["SelfMinHP"])] += 1
        final_counts[int(row["SelfFinalHP"])] += 1
    return {"min": min_counts, "final": final_counts}


def _build_summary_rows(rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    summary_rows: List[Dict[str, object]] = []
    for deck_label in ("귤", "샤를로테"):
        deck_rows = [r for r in rows if r["SelfDeck"] == deck_label]
        wins = sum(1 for r in deck_rows if r["Self W/L"] == "승")
        losses = sum(1 for r in deck_rows if r["Self W/L"] == "패")
        draws = sum(1 for r in deck_rows if r["Self W/L"] == "무")
        summary_rows.append(
            {
                "Deck": deck_label,
                "Scenarios Included": (
                    "귤 vs 귤 선/후공 + 귤 vs 샤를로테 선/후공"
                    if deck_label == "귤"
                    else "샤를로테 vs 귤 선/후공 + 샤를로테 vs 샤를로테 선/후공"
                ),
                "Match_Data rows": len(deck_rows),
                "Wins": wins,
                "Losses": losses,
                "Draws": draws,
                "Win Rate": wins / len(deck_rows) if deck_rows else 0.0,
                "Avg Self Min HP": sum(r["SelfMinHP"] for r in deck_rows) / len(deck_rows) if deck_rows else 0.0,
                "Avg Self Final HP": sum(r["SelfFinalHP"] for r in deck_rows) / len(deck_rows) if deck_rows else 0.0,
                "Max HP Check": sum(1 for r in deck_rows if r["HP_Check"] == "OK"),
            }
        )
    return summary_rows


def _build_scenario_checks(rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    checks: List[Dict[str, object]] = []
    combos = [
        ("귤", "선공", "같은 덱"),
        ("귤", "선공", "다른 덱"),
        ("귤", "후공", "같은 덱"),
        ("귤", "후공", "다른 덱"),
        ("샤를로테", "선공", "같은 덱"),
        ("샤를로테", "선공", "다른 덱"),
        ("샤를로테", "후공", "같은 덱"),
        ("샤를로테", "후공", "다른 덱"),
    ]
    for deck, side, relation in combos:
        actual = sum(1 for r in rows if r["SelfDeck"] == deck and r["SelfSide"] == side and r["Relation"] == relation)
        checks.append(
            {
                "SelfDeck": deck,
                "SelfSide": side,
                "Relation": relation,
                "Expected Rows": 100,
                "Actual Rows": actual,
                "Status": "OK" if actual == 100 else "CHECK",
            }
        )
    return checks


def _axis_text_props(font_size: int = 1000) -> RichText:
    rich = RichText()
    para_props = ParagraphProperties()
    para_props.defRPr = CharacterProperties(sz=font_size, solidFill=ColorChoice(srgbClr="000000"))
    rich.paragraphs = [Paragraph(r=[RegularTextRun(t=" ")], pPr=para_props)]
    return rich


def main() -> int:
    if not ZIP_PATH.exists():
        raise FileNotFoundError(f"zip not found: {ZIP_PATH}")
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"template xlsx not found: {TEMPLATE_PATH}")

    with zipfile.ZipFile(ZIP_PATH, "r") as zf:
        match_rows = _build_match_data(zf)

    summary_rows = _build_summary_rows(match_rows)
    scenario_checks = _build_scenario_checks(match_rows)
    hist_counts = {
        "귤": _bins_for_deck(match_rows, "귤"),
        "샤를로테": _bins_for_deck(match_rows, "샤를로테"),
    }

    wb = load_workbook(TEMPLATE_PATH)
    ws0 = wb["Match_Data"]
    match_headers = [
        "SourceFile", "MatchNo", "GameID", "SelfSide", "SelfDeck", "OppDeck", "Relation", "Result",
        "Self W/L", "Steps", "FinalTurn", "P1Deck", "P1FinalHP", "P1MinHP", "P2Deck", "P2FinalHP",
        "P2MinHP", "SelfFinalHP", "SelfMinHP", "OppFinalHP", "OppMinHP", "SelfMaxHP", "HP_Check",
    ]
    for row_idx, row in enumerate(match_rows, start=2):
        for col_idx, key in enumerate(match_headers, start=1):
            ws0.cell(row_idx, col_idx).value = row.get(key)

    ws1 = wb["Summary"]
    summary_headers = [
        "Deck", "Scenarios Included", "Match_Data rows", "Wins", "Losses", "Draws",
        "Win Rate", "Avg Self Min HP", "Avg Self Final HP", "Max HP Check",
    ]
    for row_idx, row in enumerate(summary_rows, start=2):
        for col_idx, key in enumerate(summary_headers, start=1):
            ws1.cell(row_idx, col_idx).value = row.get(key)

    ws2 = wb["Deck_Combined_400_Data"]
    ws2["B2"] = hist_counts["귤"]["min"].get(-2, 0)
    ws2["C2"] = hist_counts["귤"]["final"].get(-2, 0)
    ws2["D2"] = hist_counts["샤를로테"]["min"].get(-2, 0)
    ws2["E2"] = hist_counts["샤를로테"]["final"].get(-2, 0)
    for row_idx, hp in enumerate(HP_BINS, start=2):
        ws2.cell(row_idx, 1).value = hp
        ws2.cell(row_idx, 2).value = hist_counts["귤"]["min"].get(hp, 0)
        ws2.cell(row_idx, 3).value = hist_counts["귤"]["final"].get(hp, 0)
        ws2.cell(row_idx, 4).value = hist_counts["샤를로테"]["min"].get(hp, 0)
        ws2.cell(row_idx, 5).value = hist_counts["샤를로테"]["final"].get(hp, 0)
    ws2["G2"] = (
        "빈도값은 Match_Data의 SelfDeck/SelfMinHP/SelfFinalHP를 기준으로 Python에서 직접 집계했습니다. "
        "귤과 샤를로테 각각 4개 조합(선공/후공 × 같은/다른 덱)을 합쳐 400판으로 계산됩니다."
    )
    ws2["G2"].alignment = Alignment(wrap_text=True, vertical="top")

    ws3 = wb["Deck_Combined_400_Hist"]
    for row_idx, hp in enumerate(HP_BINS, start=2):
        ws3.cell(row_idx, 1).value = hp
        ws3.cell(row_idx, 2).value = hist_counts["귤"]["min"].get(hp, 0)
        ws3.cell(row_idx, 3).value = hist_counts["귤"]["final"].get(hp, 0)
        ws3.cell(row_idx, 4).value = hist_counts["샤를로테"]["min"].get(hp, 0)
        ws3.cell(row_idx, 5).value = hist_counts["샤를로테"]["final"].get(hp, 0)

    for chart in ws3._charts:
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.txPr = _axis_text_props(1100)
        chart.y_axis.txPr = _axis_text_props(1100)

    ws4 = wb["Scenario_Count_Check"]
    scenario_headers = ["SelfDeck", "SelfSide", "Relation", "Expected Rows", "Actual Rows", "Status"]
    for row_idx, row in enumerate(scenario_checks, start=2):
        for col_idx, key in enumerate(scenario_headers, start=1):
            ws4.cell(row_idx, col_idx).value = row.get(key)

    # keep charts/format from the template workbook, but force Excel to recalc if needed
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"saved: {OUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
